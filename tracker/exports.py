import io
from datetime import date

from django.http import HttpResponse

from .models import Beekeeper, Farm, Hive, Season, Harvest


# ── Helpers ───────────────────────────────────────────────────────────────────

MONTH_NAMES = [
    '', 'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December',
]


def _month(n):
    try:
        return MONTH_NAMES[int(n)]
    except (IndexError, TypeError, ValueError):
        return str(n)


# ── Queryset builders (with optional filters) ─────────────────────────────────

def _beekeepers_qs(params):
    qs = Beekeeper.objects.all().order_by('name')
    role = params.get('role')
    if role:
        qs = qs.filter(role=role)
    return qs


def _farms_qs(params):
    qs = Farm.objects.select_related('beekeeper').all().order_by('name')
    bk_id = params.get('beekeeper_id')
    if bk_id:
        qs = qs.filter(beekeeper_id=bk_id)
    return qs


def _hives_qs(params):
    qs = Hive.objects.select_related('farm', 'farm__beekeeper').all().order_by('farm__name', 'hive_number')
    farm_id = params.get('farm_id')
    hstatus = params.get('status')
    bk_id   = params.get('beekeeper_id')
    if farm_id: qs = qs.filter(farm_id=farm_id)
    if hstatus: qs = qs.filter(status=hstatus)
    if bk_id:   qs = qs.filter(farm__beekeeper_id=bk_id)
    return qs


def _seasons_qs(params):
    qs   = Season.objects.all().order_by('year', 'start_month')
    year = params.get('year')
    if year:
        qs = qs.filter(year=year)
    return qs


def _harvests_qs(params):
    qs = Harvest.objects.select_related(
        'hive', 'hive__farm', 'hive__farm__beekeeper', 'season'
    ).all().order_by('-harvest_date')
    farm_id   = params.get('farm_id')
    hive_id   = params.get('hive_id')
    year      = params.get('year')
    season_id = params.get('season_id')
    bk_id     = params.get('beekeeper_id')
    if farm_id:   qs = qs.filter(hive__farm_id=farm_id)
    if hive_id:   qs = qs.filter(hive_id=hive_id)
    if year:      qs = qs.filter(harvest_date__year=year)
    if season_id: qs = qs.filter(season_id=season_id)
    if bk_id:     qs = qs.filter(hive__farm__beekeeper_id=bk_id)
    return qs


# ── Row factories ─────────────────────────────────────────────────────────────

RESOURCE_HEADERS = {
    'beekeepers': ['ID', 'Name', 'Email', 'Role', 'Created At'],
    'farms':      ['ID', 'Name', 'Location', 'Established Date', 'Beekeeper'],
    'hives':      ['ID', 'Hive Number', 'Type', 'Install Date', 'Queen Status', 'Status', 'Farm', 'Farm Location', 'Beekeeper'],
    'seasons':    ['ID', 'Name', 'Year', 'Start Month', 'End Month'],
    'harvests':   ['ID', 'Harvest Date', 'Yield (kg)', 'Notes', 'Hive Number', 'Farm', 'Farm Location', 'Season', 'Season Year', 'Beekeeper'],
}


def _beekeepers_rows(qs):
    for b in qs:
        yield [
            b.id, b.name, b.email, b.get_role_display(),
            b.created_at.strftime('%Y-%m-%d %H:%M') if b.created_at else '',
        ]


def _farms_rows(qs):
    for f in qs:
        yield [
            f.id, f.name, f.location,
            f.established_date.strftime('%Y-%m-%d') if f.established_date else '',
            f.beekeeper.name,
        ]


def _hives_rows(qs):
    for h in qs:
        yield [
            h.id, h.hive_number, h.get_hive_type_display(),
            h.install_date.strftime('%Y-%m-%d') if h.install_date else '',
            h.get_queen_status_display(), h.get_status_display(),
            h.farm.name, h.farm.location, h.farm.beekeeper.name,
        ]


def _seasons_rows(qs):
    for s in qs:
        yield [s.id, s.name, s.year, _month(s.start_month), _month(s.end_month)]


def _harvests_rows(qs):
    for h in qs:
        yield [
            h.id,
            h.harvest_date.strftime('%Y-%m-%d') if h.harvest_date else '',
            round(h.yield_kg, 2),
            h.notes or '',
            h.hive.hive_number,
            h.hive.farm.name,
            h.hive.farm.location,
            h.season.name if h.season else '',
            h.season.year if h.season else '',
            h.hive.farm.beekeeper.name,
        ]


RESOURCE_MAP = {
    'beekeepers': (_beekeepers_qs, _beekeepers_rows),
    'farms':      (_farms_qs,      _farms_rows),
    'hives':      (_hives_qs,      _hives_rows),
    'seasons':    (_seasons_qs,    _seasons_rows),
    'harvests':   (_harvests_qs,   _harvests_rows),
}


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(resource, params):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    qs_fn, rows_fn = RESOURCE_MAP[resource]
    qs = qs_fn(params)
    headers = RESOURCE_HEADERS[resource]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = resource.capitalize()

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill(fill_type='solid', fgColor='F5A623')
    center_align = Alignment(horizontal='center')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    for row_idx, row in enumerate(rows_fn(qs), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{resource}_{date.today().isoformat()}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── PDF export ────────────────────────────────────────────────────────────────

def export_pdf(resource, params):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    qs_fn, rows_fn = RESOURCE_MAP[resource]
    qs = qs_fn(params)
    headers = RESOURCE_HEADERS[resource]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1 * cm, leftMargin=1 * cm,
        topMargin=1.5 * cm, bottomMargin=1 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.textColor = colors.HexColor('#5C3D11')

    HONEY_AMBER = colors.HexColor('#F5A623')
    LIGHT_AMBER = colors.HexColor('#FFF3D6')

    data = [headers] + list(rows_fn(qs))

    # Wrap long strings so they don't overflow
    normal = styles['Normal']
    normal.fontSize = 7
    wrapped_data = []
    for row in data:
        wrapped_data.append([Paragraph(str(cell), normal) for cell in row])

    col_count  = len(headers)
    page_width = landscape(A4)[0] - 2 * cm
    col_width  = page_width / col_count

    table = Table(wrapped_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0),  HONEY_AMBER),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_AMBER]),
        ('FONTSIZE',    (0, 1), (-1, -1), 7),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('VALIGN',      (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING',  (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    title = Paragraph(
        f"<b>Bee Tracker — {resource.capitalize()}</b> &nbsp;"
        f"<font size='8' color='grey'>Exported {date.today().isoformat()}</font>",
        title_style,
    )
    elements = [title, Spacer(1, 0.4 * cm), table]
    doc.build(elements)

    buf.seek(0)
    filename = f"{resource}_{date.today().isoformat()}.pdf"
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
